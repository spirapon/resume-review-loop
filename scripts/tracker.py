#!/usr/bin/env python3
"""Update application_tracker.xlsx.

Usage:
  tracker.py add <app_folder>
  tracker.py update <app_folder> --status <s> [--reason "..."]
  tracker.py score
  tracker.py worth
  tracker.py tokens <app_folder> [--count N]

Statuses: rejected_visa, rejected_language, screened, in_progress, done,
stopped_best_effort, skipped.

The tracker is hand-edited: columns may be deleted or reordered, and the ones
in USER_COLUMNS belong to the user alone — see put().
"""
import argparse
import datetime
import json
import re
import sys
from pathlib import Path

import openpyxl
import yaml

from ats_check import load_keywords, normalize
from config import resume_prefix


def sanitize(name):
    """Same rule as render.py so the tracker matches the real file name."""
    name = str(name).replace("/", "-").replace(" ", "_")
    return re.sub(r"[^A-Za-z0-9._-]", "", name)

PROJECT = Path(__file__).resolve().parent.parent
TRACKER = PROJECT / "application_tracker.xlsx"
LOCK = PROJECT / "~$application_tracker.xlsx"
MASTER = PROJECT / "0_master_resume" / "master.yaml"

STATUSES = {"rejected_visa", "rejected_language", "screened", "in_progress",
            "done", "stopped_best_effort", "skipped"}


def check_lock():
    if LOCK.exists():
        print("ERROR: application_tracker.xlsx is open in Excel. "
              "Close it and run again.")
        sys.exit(1)


def headers(ws):
    return {str(c.value): i + 1 for i, c in enumerate(ws[1]) if c.value}


# Columns that belong to the user. No script may write, create or delete these
# — they are filled in by hand and a script touching them would destroy work.
USER_COLUMNS = {"CodexScore", "applyDate"}


def put(ws, h, row, column, value):
    """Write a cell only if that column exists in the sheet.

    The tracker is a spreadsheet a human edits: columns get deleted and
    reordered. Optional columns are skipped rather than recreated, so a
    deleted column stays deleted. Writing to a USER_COLUMNS entry is a bug,
    so it raises rather than failing quietly.
    """
    if column in USER_COLUMNS:
        raise AssertionError(f"{column} is the user's column; scripts must never write it")
    if column in h:
        ws.cell(row=row, column=h[column], value=value)


def match_score(folder):
    """Pre-screen score: % of the JD's keywords already in master.yaml.
    Higher = your background already covers the JD = easier to tailor.
    Returns an int 0-100, or None if inputs are missing."""
    kw_path = folder / "keywords.txt"
    if not kw_path.exists() or not MASTER.exists():
        return None
    buckets = load_keywords(kw_path)
    text = normalize(MASTER.read_text())
    hit = total = 0
    for bucket in ("P1", "P2", "P3"):
        words = buckets[bucket]
        total += len(words)
        hit += sum(1 for w in words if normalize(w).strip() in text)
    return round(100 * hit / total) if total else None


def worth_score(folder):
    """Read the 'WORTH: n' line the fit-scorer agent wrote to worth_report.txt.

    This only transcribes — the judgement lives in the agent, so there is
    nothing here to recompute. Returns an int 0-100, or None if unscored.
    """
    report = folder / "worth_report.txt"
    if not report.exists():
        return None
    m = re.search(r"(?im)^WORTH:\s*(\d{1,3})\s*$", report.read_text())
    if not m:
        return None
    return min(int(m.group(1)), 100)


def find_row(ws, col, folder_name):
    for row in range(2, ws.max_row + 1):
        if str(ws.cell(row=row, column=col).value) == folder_name:
            return row
    return None


def cmd_add(folder, status="in_progress"):
    analysis = yaml.safe_load((folder / "analysis.yaml").read_text())
    wb = openpyxl.load_workbook(TRACKER)
    ws = wb.active
    h = headers(ws)

    if find_row(ws, h["folder name"], folder.name):
        print(f"already tracked: {folder.name}")
        return

    row = ws.max_row + 1
    nums = [ws.cell(row=r, column=h["NO."]).value for r in range(2, ws.max_row + 1)]
    nums = [n for n in nums if isinstance(n, int)]
    put(ws, h, row, "NO.", max(nums) + 1 if nums else 1)
    put(ws, h, row, "date", int(datetime.date.today().strftime("%Y%m%d")))
    put(ws, h, row, "source", analysis.get("source", ""))
    put(ws, h, row, "position", analysis.get("position", ""))
    put(ws, h, row, "company", analysis.get("company", ""))
    put(ws, h, row, "location", analysis.get("location", ""))
    put(ws, h, row, "requireLanguage", analysis.get("language_requirement", ""))
    put(ws, h, row, "requireVISA", analysis.get("visa_sponsorship", ""))
    put(ws, h, row, "link", analysis.get("link", ""))
    put(ws, h, row, "resume name",
        f"{resume_prefix()}_resume_{sanitize(analysis.get('company', ''))}"
        f"_{sanitize(analysis.get('position', ''))}")
    put(ws, h, row, "status", status)
    ws.cell(row=row, column=h["folder name"], value=folder.name)  # required: the join key

    score = match_score(folder)
    if score is not None:
        if "matchScore" not in h:  # add the column the first time we need it
            h["matchScore"] = ws.max_column + 1
            ws.cell(row=1, column=h["matchScore"], value="matchScore")
        ws.cell(row=row, column=h["matchScore"], value=score)

    worth = worth_score(folder)
    if worth is not None:
        if "worthScore" not in h:
            h["worthScore"] = ws.max_column + 1
            ws.cell(row=1, column=h["worthScore"], value="worthScore")
        ws.cell(row=row, column=h["worthScore"], value=worth)

    wb.save(TRACKER)
    print(f"added: {folder.name} (matchScore={score}, worthScore={worth})")


def cmd_update(folder, status, reason):
    wb = openpyxl.load_workbook(TRACKER)
    ws = wb.active
    h = headers(ws)
    row = find_row(ws, h["folder name"], folder.name)
    if row is None:
        print(f"ERROR: {folder.name} not found in tracker (run 'add' first)")
        sys.exit(1)
    put(ws, h, row, "status", status)
    if reason:
        put(ws, h, row, "reason", reason)
    # Mark the resume as finished with a checkmark when status is 'done'.
    put(ws, h, row, "resumeDone", "✓" if status == "done" else "")

    wb.save(TRACKER)
    print(f"updated: {folder.name} -> {status}")


def cmd_score():
    """(Re)compute matchScore for every tracked row that has a folder on disk."""
    apps = PROJECT / "2_applications"
    wb = openpyxl.load_workbook(TRACKER)
    ws = wb.active
    h = headers(ws)
    if "matchScore" not in h:
        h["matchScore"] = ws.max_column + 1
        ws.cell(row=1, column=h["matchScore"], value="matchScore")
    n = 0
    for row in range(2, ws.max_row + 1):
        name = ws.cell(row=row, column=h["folder name"]).value
        if not name:
            continue
        score = match_score(apps / str(name))
        if score is not None:
            ws.cell(row=row, column=h["matchScore"], value=score)
            n += 1
    wb.save(TRACKER)
    print(f"scored: {n} rows")


def cmd_worth():
    """Copy each folder's worth_report.txt score into the worthScore column.

    Only folders the fit-scorer agent has already scored are updated; this
    command never invents a score for an unscored folder.
    """
    apps = PROJECT / "2_applications"
    wb = openpyxl.load_workbook(TRACKER)
    ws = wb.active
    h = headers(ws)
    if "worthScore" not in h:
        h["worthScore"] = ws.max_column + 1
        ws.cell(row=1, column=h["worthScore"], value="worthScore")
    n = 0
    for row in range(2, ws.max_row + 1):
        name = ws.cell(row=row, column=h["folder name"]).value
        if not name:
            continue
        worth = worth_score(apps / str(name))
        if worth is not None:
            ws.cell(row=row, column=h["worthScore"], value=worth)
            n += 1
    wb.save(TRACKER)
    print(f"worth-scored: {n} rows")


def cmd_tokens(folder, count):
    """Record how many tokens the whole loop spent on this application.

    There is nothing on disk to recompute this from — the orchestrator is the
    only thing that sees each subagent's usage — so the count is passed in and
    also stashed in status.json, which makes it survive a re-run of this
    command with no --count.
    """
    state = folder / "status.json"
    if count is None:
        if not state.exists():
            print(f"ERROR: no --count given and no status.json in {folder.name}")
            sys.exit(2)
        count = json.loads(state.read_text()).get("tokens_used")
        if count is None:
            print("ERROR: no --count given and status.json has no tokens_used")
            sys.exit(2)
    elif state.exists():
        s = json.loads(state.read_text())
        s["tokens_used"] = count
        state.write_text(json.dumps(s, indent=1) + "\n")

    wb = openpyxl.load_workbook(TRACKER)
    ws = wb.active
    h = headers(ws)
    row = find_row(ws, h["folder name"], folder.name)
    if row is None:
        print(f"ERROR: {folder.name} not found in tracker (run 'add' first)")
        sys.exit(1)
    if "tokensUsed" not in h:  # add the column the first time we need it
        h["tokensUsed"] = ws.max_column + 1
        ws.cell(row=1, column=h["tokensUsed"], value="tokensUsed")
    ws.cell(row=row, column=h["tokensUsed"], value=count)
    wb.save(TRACKER)
    print(f"tokens: {folder.name} -> {count}")


def main():
    p = argparse.ArgumentParser()
    p.add_argument("command", choices=["add", "update", "score", "worth", "tokens"])
    p.add_argument("app_folder", nargs="?")
    p.add_argument("--status", choices=sorted(STATUSES))
    p.add_argument("--reason", default="")
    p.add_argument("--count", type=int, help="tokens spent on this application")
    args = p.parse_args()

    check_lock()
    if args.command == "score":
        cmd_score()
        return
    if args.command == "worth":
        cmd_worth()
        return
    if not args.app_folder:
        print(f"ERROR: {args.command} needs an app_folder")
        sys.exit(2)
    folder = Path(args.app_folder).resolve()
    if args.command == "tokens":
        cmd_tokens(folder, args.count)
        return
    if args.command == "add":
        cmd_add(folder, status=args.status or "in_progress")
    else:
        if not args.status:
            print("ERROR: update needs --status")
            sys.exit(2)
        cmd_update(folder, args.status, args.reason)


if __name__ == "__main__":
    main()
